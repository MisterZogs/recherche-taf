"""Dédoublonnage du classeur offres_emploi.xlsx (02/09/2026).

Repéré par Gaëtan : les lignes 6 et 8 de l'onglet SIRH portaient la même offre
(Solution Consultant HCM chez Workday, même lien Workday CXS). Le balayage
complet a montré que le cas était loin d'être isolé.

Principe : une page d'offre individuelle ne peut correspondre qu'à un seul
poste. Tout lien individuel partagé par plusieurs lignes est donc un doublon,
et les lignes sont fusionnées en une seule.

Ne sont PAS traités ici (bug distinct, déjà documenté dans CLAUDE.md) :
les liens génériques (pages catégorie/recherche collées à la place du lien
individuel), qui sont seulement listés en fin d'exécution.
"""

import re
import unicodedata
from difflib import SequenceMatcher

import openpyxl

from add_offre import _capture_cell, _col_index, _ecrire_onglet, _restore_cell

FICHIER = 'offres_emploi.xlsx'

ONGLETS = ['Offres SIRH', 'Pays Basque', 'Offres USA', 'Offres PM',
           'Offres CSM', 'Offres IA', 'NoRemote', 'Fait']
ACTIFS = {'Offres SIRH', 'Pays Basque', 'Offres USA', 'Offres PM',
          'Offres CSM', 'Offres IA'}

# Rang de conservation : plus le rang est bas, plus la ligne fait autorité.
# Fait = offre déjà traitée, son statut est à jour.
# NoRemote = porte en général l'information de télétravail vérifiée en fiche
# (hybride/partiel/présentiel), là où l'onglet métier se contente de « n.p. ».
RANG_ONGLET = {'Fait': 0, 'NoRemote': 1}

PRIO_ORDRE = {'⭐⭐⭐⭐⭐': 0, '⭐⭐⭐⭐': 1, '⭐⭐⭐': 2, '⭐⭐': 3, '⭐': 4}

# Pages catégorie / recherche : plusieurs offres légitimement derrière une même
# URL. À ne pas fusionner.
GENERIQUES = [
    re.compile(r'linkedin\.com/jobs/[^/]*emplois', re.I),
    re.compile(r'indeed\.fr/q-', re.I),
    re.compile(r'free-work\.com/fr/tech-it/jobs/', re.I),
    re.compile(r'convictionsrh\.com/offre-emploi', re.I),
    re.compile(r'careers\.ey\.com/ey/search', re.I),
    re.compile(r'whitehallresources\.com/jobs/\?', re.I),
    re.compile(r'sqorus\.com/carrieres', re.I),
    re.compile(r'freelance-informatique\.fr/mission-[a-z-]+-\d{1,4}$', re.I),
]

# Cas tranchés à la main, en vérifiant la source. Le lien pointe vers la ligne
# (onglet, poste normalisé) à conserver, contre la règle de rang ci-dessus.
GARDER_EXPLICITE = {
    # « Le poste est assuré à distance sauf pour le tournage » (fiche werecruit) :
    # le télétravail est bien la règle, la ligne NoRemote était fausse.
    'https://careers.werecruit.io/fr/walter-learning/offres/'
    'formateur---intelligence-artificielle-generative-3cde70': 'Offres IA',
}

# Lien vérifié trompeur : l'API WTTJ renvoie pour ce slug une offre
# d'alternance « Bras droit Responsable BU AI for Finance », sans rapport avec
# le poste enregistré. Lien vidé et raison documentée, conformément à la règle
# absolue de CLAUDE.md sur les liens qui ne correspondent pas à l'offre.
LIEN_TROMPEUR = {
    'https://www.welcometothejungle.com/fr/companies/mister-ia/jobs/'
    'chef-de-projet-bu-conseil-h-f-cdi-paris-8e-asap_paris':
        "Lien retiré le 02/09/2026 : le slug WTTJ renvoie en réalité une "
        "alternance « Bras droit Responsable BU AI for Finance », sans rapport "
        "avec ce poste. URL individuelle à retrouver avant de candidater.",
}

ANONYME = re.compile(
    r'^\s*$|n\.?/?c\b|anonym|non\s+(communiqu|divulgu|pr[ée]cis|sp[ée]cifi)|'
    r'confidentiel|^client\b', re.I)

# Noms de plateformes : la colonne Entreprise porte souvent le site d'où vient
# l'annonce plutôt que l'employeur. Deux lignes qui diffèrent seulement sur ce
# point décrivent la même offre.
PLATEFORMES = re.compile(
    r'\b(via|freelance[- ]informatique|mission[- ]freelances?|free[- ]work|'
    r'workdispo|eursap|movement group|michael page|welcome to the jungle|'
    r'linkedin|indeed|jobgether|wizbii|whitehall|hays)\b', re.I)


def norm(s):
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', ' ', s.lower()).strip()


def anonyme(v):
    return bool(ANONYME.search(str(v or '')))


def nom_employeur(v):
    """Nom d'employeur exploitable, ou '' si la cellule ne porte qu'une source."""
    if anonyme(v):
        return ''
    s = re.sub(r'\([^)]*\)', ' ', str(v or ''))      # « (via Michael Page) »
    s = PLATEFORMES.sub(' ', s)
    return norm(s)


def memes_employeurs(noms):
    """Variantes d'un même employeur ? « Deloitte » vs « Deloitte France »."""
    ref = max(noms, key=len)
    for n in noms:
        court, long_ = sorted([n, ref], key=len)
        if court in long_ or SequenceMatcher(None, n, ref).ratio() >= 0.8:
            continue
        return False
    return True


def est_generique(lien):
    return any(p.search(lien) for p in GENERIQUES)


def titres_compatibles(titres):
    """Deux intitulés décrivent-ils le même poste ?"""
    ref = max(titres, key=len)
    for t in titres:
        court, long_ = sorted([t, ref], key=len)
        if court and court in long_:
            continue
        if SequenceMatcher(None, t, ref).ratio() < 0.6:
            return False
    return True


def main():
    wb = openpyxl.load_workbook(FICHIER)

    # 1. Capture de toutes les lignes, styles compris.
    lignes = {}          # (onglet, index) -> row_data
    idx_col = {}
    for sn in ONGLETS:
        ws = wb[sn]
        idx_col[sn] = {h.value: i for i, h in enumerate(ws[1])}
        for r in range(2, ws.max_row + 1):
            row = [_capture_cell(c) for c in ws[r]]
            if all(d['value'] is None for d in row):
                continue
            lignes[(sn, r)] = row

    def champ(cle, nom):
        i = idx_col[cle[0]].get(nom)
        return lignes[cle][i]['value'] if i is not None else None

    # 2. Regroupement par lien.
    groupes = {}
    for cle in lignes:
        lien = champ(cle, 'Lien')
        if lien and str(lien).strip():
            groupes.setdefault(str(lien).strip().rstrip('/'), []).append(cle)

    a_supprimer = set()
    fusions, signales = [], []

    for lien, cles in sorted(groupes.items()):
        if len(cles) < 2:
            continue

        entreprises = {norm(champ(c, 'Entreprise')) for c in cles
                       if not anonyme(champ(c, 'Entreprise'))}
        titres = [norm(champ(c, 'Poste')) for c in cles]

        if est_generique(lien) or len(entreprises) > 1 or not titres_compatibles(titres):
            signales.append((lien, cles))
            continue

        # 3. Choix de la ligne conservée.
        onglet_force = GARDER_EXPLICITE.get(lien)

        def cle_tri(c):
            return (0 if c[0] == onglet_force else 1,
                    RANG_ONGLET.get(c[0], 2),
                    PRIO_ORDRE.get(champ(c, 'Priorité'), 9),
                    -len(str(champ(c, 'Fit / Notes') or '')),
                    c[1])

        gardee, *autres = sorted(cles, key=cle_tri)
        idx = idx_col[gardee[0]]

        # 4. Fusion : on récupère des lignes supprimées ce que la ligne
        #    conservée n'a pas.
        for autre in autres:
            for nom, i in idx.items():
                j = idx_col[autre[0]].get(nom)
                if j is None:
                    continue
                mien, sien = lignes[gardee][i]['value'], lignes[autre][j]['value']
                if sien in (None, ''):
                    continue
                if nom == 'Priorité':
                    if PRIO_ORDRE.get(sien, 9) < PRIO_ORDRE.get(mien, 9):
                        lignes[gardee][i]['value'] = sien
                elif nom == 'Fit / Notes':
                    if len(str(sien)) > len(str(mien or '')):
                        lignes[gardee][i]['value'] = sien
                elif nom == 'Entreprise':
                    if anonyme(mien) and not anonyme(sien):
                        lignes[gardee][i]['value'] = sien
                elif nom in ('Fait', 'Remote'):
                    pass                      # la ligne conservée fait foi
                elif mien in (None, ''):
                    lignes[gardee][i]['value'] = sien

        if lien in LIEN_TROMPEUR:
            lignes[gardee][idx['Lien']]['value'] = None
            lignes[gardee][idx['Lien']]['hyperlink'] = None
            note = lignes[gardee][idx['Fit / Notes']]['value'] or ''
            lignes[gardee][idx['Fit / Notes']]['value'] = (
                f"{note} {LIEN_TROMPEUR[lien]}".strip())

        a_supprimer.update(autres)
        fusions.append((lien, gardee, autres))

    # 5. Réécriture.
    print(f"{len(fusions)} lien(s) dédoublonné(s), "
          f"{len(a_supprimer)} ligne(s) supprimée(s)\n")
    for lien, gardee, autres in fusions:
        perdus = ', '.join(f'{s} L{r}' for s, r in autres)
        print(f"  garde {gardee[0]} L{gardee[1]:<4} | supprime {perdus}")
        print(f"        {lien}")

    for sn in ONGLETS:
        rows = [row for (s, r), row in sorted(lignes.items())
                if s == sn and (s, r) not in a_supprimer]
        if sn == 'Fait':
            # Archive : on préserve l'ordre historique, comme ajouter_offres().
            ws = wb[sn]
            ws.delete_rows(2, ws.max_row)
            for i, row_data in enumerate(rows):
                for j, cell_data in enumerate(row_data):
                    _restore_cell(ws.cell(row=i + 2, column=j + 1), cell_data)
            print(f"  {sn} : {len(rows)} offres")
        else:
            _ecrire_onglet(wb[sn], rows, verbose=True)

    wb.save(FICHIER)

    print(f"\n{len(signales)} lien(s) partagé(s) NON fusionné(s) "
          f"(page catégorie ou offres réellement distinctes) :")
    for lien, cles in signales:
        actifs = sum(1 for c in cles if c[0] in ACTIFS)
        print(f"  {len(cles)} lignes ({actifs} en onglet actif) : {lien}")


if __name__ == '__main__':
    main()
