"""
Utilitaire pour ajouter des offres dans offres_emploi.xlsx.

- Les offres CSM (Customer Success) vont dans l'onglet "Offres CSM".
- Les offres IA (formateur IA, IA x SIRH, IA x RH) vont dans "Offres IA".
- Les offres Product Manager / Product Owner vont dans "Offres PM",
  sauf si l'intitulé porte aussi un marqueur SIRH/SAP (elles restent alors en SIRH).
- Les offres SIRH/SAP vont dans l'onglet "Offres SIRH".
- FILTRE PRIORITAIRE : une offre qui exclut explicitement le télétravail total
  (hybride, partiel, présentiel) va dans "NoRemote", quel que soit le métier.
  Une information manquante ne disqualifie plus : l'offre reste dans son
  onglet métier.
- Avant chaque ajout, les lignes marquées "x" dans la colonne Fait
  sont déplacées vers l'onglet "Fait".
"""

import re

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

FICHIER = "offres_emploi.xlsx"

COLS = [
    'Priorité', 'Statut', 'Fait', 'Poste', 'Entreprise', 'Source', 'Lien', 'Contrat',
    'Localisation', 'Remote', 'Salaire / TJM', 'Durée mission',
    'Fit / Notes', 'CV à envoyer', 'Prétention'
]

CSM_KEYWORDS = ['Customer Success', 'Client Success', 'CSM']
IA_KEYWORDS  = ['Formateur IA', 'Formation IA', 'IA générative', 'IA x SIRH', 'IA x RH',
                 'Intelligence Artificielle', 'AI Trainer', 'GenAI', 'LLM', 'Prompt',
                 'Plateforme IA', 'AI Platform', 'Projet IA', 'PMO IA']
PM_KEYWORDS  = ['Product Manager', 'Product Owner', 'Product Lead', 'Head of Product',
                 'Director of Product', 'VP Product', 'Chef de Produit', 'Responsable Produit',
                 'Product Marketing Manager', 'Product Builder', 'Proxy PO']

# Métiers d'avant-vente et de gestion de compte technique : même famille que le CSM,
# donc rangés dans "Offres CSM".
PRESALES_KEYWORDS = ['Technical Account Manager', 'Solutions Engineer', 'Solution Engineer',
                      'Sales Engineer', 'Solutions Consultant', 'Solution Consultant',
                      'Pre-Sales', 'Presales', 'Pre-sales', 'Présales', 'Avant-vente',
                      'Account Manager', 'Solutions Advocate', 'Solution Advocate',
                      'Solutions Architect', 'Solution Architect', 'Solution Advisor',
                      'Solutions Sales Executive']

# Une offre Product Manager ou avant-vente dont l'intitulé porte aussi un marqueur
# SIRH/SAP reste dans "Offres SIRH" : le métier SIRH prime sur le titre.
SIRH_OVERRIDE = ['SIRH', 'HRIS', 'SAP', 'SuccessFactors', 'Workday', 'HCM',
                  'HR Access', 'Paie', 'Payroll', 'RH', 'HXM', 'HR ']

COLORS = {
    '⭐⭐⭐⭐⭐': '00FF0000',
    '⭐⭐⭐⭐':   '00FF8C00',
    '⭐⭐⭐':     '00FFD700',
    '⭐⭐':       '0070AD47',
    '⭐':         '00969696',
}

PRIORITY_ORDER = {
    '⭐⭐⭐⭐⭐': 0,
    '⭐⭐⭐⭐':   1,
    '⭐⭐⭐':     2,
    '⭐⭐':       3,
    '⭐':         4,
}

# Ordre de tri par statut. Les offres traitées (postulé, refusé) remontent en
# tête ; les offres mortes descendent tout en bas, sous les offres actives.
# Tout le reste (cellule vide, "À postuler", "À vérifier"...) occupe le rang 2,
# qui est aussi la valeur par défaut dans sort_key.
STATUS_ORDER = {
    'Postulé':  0,
    'Refusé':   1,
    'Expiré':   3,
    'Expirée':  3,
}
STATUS_DEFAUT = 2


# ── Filtre télétravail ──────────────────────────────────────────────────────
# Règle révisée le 18/08/2026 : une information manquante ne disqualifie plus
# une offre. Partent dans "NoRemote" les seules offres qui excluent
# explicitement le télétravail total, c'est-à-dire l'hybride, le partiel et le
# présentiel confirmés. Une colonne Remote vide, un "n.p." ou un "à vérifier"
# laisse l'offre dans son onglet métier, à charge de clarifier ensuite.
#
# Marqueurs qui disqualifient, même si "remote" ou "oui" apparaît ailleurs
# (ex. "Hybride (3j remote + 2j sur site)").
_REMOTE_NON = re.compile(
    r'(hybrid|partiel|pr[ée]sentiel|sur site|on\s*-?\s*site|\d\s*j\b|\d\s*jours'
    r'|^non$|^no$)', re.I)


def accepte_remote(valeur) -> bool:
    """Faux seulement si la valeur exclut explicitement le télétravail total."""
    s = str(valeur).strip() if valeur is not None else ''
    return not _REMOTE_NON.search(s)


def _hors_sirh(poste: str) -> bool:
    """Vrai si l'intitulé ne porte aucun marqueur SIRH/SAP."""
    p = poste.lower()
    return not any(kw.lower() in p for kw in SIRH_OVERRIDE)


def _is_csm(poste: str) -> bool:
    if any(kw in poste for kw in CSM_KEYWORDS):
        return True
    # Avant-vente et TAM : rattachés au CSM, sauf s'il s'agit d'un poste SIRH/SAP,
    # qui reste dans "Offres SIRH".
    if any(kw.lower() in poste.lower() for kw in PRESALES_KEYWORDS):
        return _hors_sirh(poste)
    return False


def _is_ia(poste: str) -> bool:
    if any(kw.lower() in poste.lower() for kw in IA_KEYWORDS):
        return True
    # Attrape les intitulés où "IA" / "AI" est un mot isolé : "Expert IA",
    # "Solutions IA", "PMO - IA", "Méthode & IA". Sensible à la casse pour
    # éviter de matcher "ai" dans "j'ai", "vrai", "media".
    return bool(re.search(r'\b(IA|AI)\b', poste))


def _is_pm(poste: str) -> bool:
    if not any(kw.lower() in poste.lower() for kw in PM_KEYWORDS):
        return False
    return _hors_sirh(poste)


def _capture_cell(cell):
    return {
        'value':         cell.value,
        'data_type':     cell.data_type,
        'font':          copy(cell.font)       if cell.has_style else None,
        'border':        copy(cell.border)     if cell.has_style else None,
        'fill':          copy(cell.fill)       if cell.has_style else None,
        'number_format': cell.number_format    if cell.has_style else None,
        'protection':    copy(cell.protection) if cell.has_style else None,
        'alignment':     copy(cell.alignment)  if cell.has_style else None,
        'hyperlink':     cell.hyperlink,
    }


def _restore_cell(cell, data):
    cell.value = data['value']
    if data['font']:          cell.font          = data['font']
    if data['border']:        cell.border        = data['border']
    if data['fill']:          cell.fill          = data['fill']
    if data['number_format']: cell.number_format = data['number_format']
    if data['protection']:    cell.protection    = data['protection']
    if data['alignment']:     cell.alignment     = data['alignment']
    if data['hyperlink']:     cell.hyperlink     = data['hyperlink']


def _clear_cell(cell):
    cell.value = None


def _col_index(ws, name):
    for i, cell in enumerate(ws[1]):
        if cell.value == name:
            return i
    raise ValueError(f"Colonne '{name}' introuvable dans {ws.title}")


def _liens_existants(wb) -> set:
    """Tous les liens déjà présents dans le classeur, tous onglets confondus (y compris Fait)."""
    liens = set()
    for ws in wb.worksheets:
        try:
            lien_idx = _col_index(ws, 'Lien')
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2):
            v = row[lien_idx].value
            if v:
                liens.add(str(v).strip())
    return liens


def _archiver_faits(ws_src, ws_fait, fait_idx, verbose):
    """Déplace les lignes 'x' de ws_src vers ws_fait. Retourne les lignes conservées."""
    kept = []
    archived = []
    for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row):
        row_data = [_capture_cell(c) for c in row]
        # Ignorer les lignes entièrement vides
        if all(d['value'] is None for d in row_data):
            continue
        if str(row_data[fait_idx]['value'] or '').strip().lower() == 'x':
            archived.append(row_data)
        else:
            kept.append(row_data)

    if archived:
        if ws_fait.max_row == 1 and ws_fait.cell(1, 1).value is None:
            for j, cell in enumerate(ws_src[1]):
                ws_fait.cell(row=1, column=j + 1).value = cell.value
        next_row = ws_fait.max_row + 1
        for row_data in archived:
            for j, cell_data in enumerate(row_data):
                _restore_cell(ws_fait.cell(row=next_row, column=j + 1), cell_data)
            next_row += 1
        if verbose:
            print(f"  Archivé ({ws_src.title}) → Fait : {len(archived)} ligne(s)")

    return kept


def _ecrire_onglet(ws, rows, verbose):
    """Trie et réécrit les lignes dans ws, efface le surplus."""
    prio_idx   = _col_index(ws, 'Priorité')
    status_idx = _col_index(ws, 'Statut')

    def sort_key(r):
        status   = r[status_idx]['value'] or ''
        priority = r[prio_idx]['value']   or ''
        return (STATUS_ORDER.get(status, STATUS_DEFAUT),
                PRIORITY_ORDER.get(priority, 99))

    rows.sort(key=sort_key)

    # Suppression physique des anciennes lignes plutôt qu'un simple vidage des
    # valeurs : vider les cellules laissait des lignes fantômes (une cellule
    # isolée survivait et réapparaissait comme une offre sans intitulé).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for i, row_data in enumerate(rows):
        for j, cell_data in enumerate(row_data):
            _restore_cell(ws.cell(row=i + 2, column=j + 1), cell_data)

    if verbose:
        print(f"  {ws.title} : {len(rows)} offres")


def _nouvelle_ligne(offre: dict) -> list:
    row_data = []
    for col in COLS:
        val = offre.get(col)
        cell_data = {
            'value': val, 'data_type': None, 'font': None,
            'border': None, 'fill': None, 'number_format': None,
            'protection': None, 'alignment': None, 'hyperlink': None,
        }
        if col == 'Priorité' and val in COLORS:
            cell_data['fill'] = PatternFill(fill_type='solid', fgColor=COLORS[val])
        row_data.append(cell_data)
    return row_data


def ajouter_offres(offres: list[dict], verbose=True):
    """
    1. Archive les lignes 'x' (colonne Fait) de chaque onglet vers Fait.
    2. Route les nouvelles offres : CSM → Offres CSM, IA → Offres IA, autres → Offres SIRH.
    3. Trie chaque onglet.
    """
    wb = openpyxl.load_workbook(FICHIER)
    ws_sirh = wb['Offres SIRH']
    ws_csm  = wb['Offres CSM']
    ws_ia   = wb['Offres IA']
    ws_pm   = wb['Offres PM']
    ws_nore = wb['NoRemote']
    ws_fait = wb['Fait']

    fait_idx = _col_index(ws_sirh, 'Fait')

    liens_connus = _liens_existants(wb)

    if verbose:
        print("── Archivage ──")

    rows_sirh = _archiver_faits(ws_sirh, ws_fait, fait_idx, verbose)
    rows_csm  = _archiver_faits(ws_csm,  ws_fait, fait_idx, verbose)
    rows_ia   = _archiver_faits(ws_ia,   ws_fait, fait_idx, verbose)
    rows_pm   = _archiver_faits(ws_pm,   ws_fait, fait_idx, verbose)
    rows_nore = _archiver_faits(ws_nore, ws_fait, fait_idx, verbose)

    if verbose:
        print("── Ajout ──")

    for offre in offres:
        poste = offre.get('Poste', '')
        lien = str(offre.get('Lien') or '').strip()
        if lien and lien in liens_connus:
            if verbose:
                print(f"= [doublon ignoré] {poste} | {offre.get('Entreprise')} | {lien}")
            continue
        if lien:
            liens_connus.add(lien)
        ligne = _nouvelle_ligne(offre)
        # Sans télétravail confirmé, l'offre va dans NoRemote quel que soit le métier.
        if not accepte_remote(offre.get('Remote')):
            rows_nore.append(ligne)
            if verbose:
                print(f"+ [NoRemote] {offre.get('Priorité')} | {poste} | {offre.get('Entreprise')}")
            continue
        if _is_ia(poste):
            rows_ia.append(ligne)
            target = 'Offres IA'
        elif _is_csm(poste):
            rows_csm.append(ligne)
            target = 'Offres CSM'
        elif _is_pm(poste):
            rows_pm.append(ligne)
            target = 'Offres PM'
        else:
            rows_sirh.append(ligne)
            target = 'Offres SIRH'
        if verbose:
            print(f"+ [{target}] {offre.get('Priorité')} | {poste} | {offre.get('Entreprise')}")

    if verbose:
        print("── Tri & sauvegarde ──")

    _ecrire_onglet(ws_sirh, rows_sirh, verbose)
    _ecrire_onglet(ws_csm,  rows_csm,  verbose)
    _ecrire_onglet(ws_ia,   rows_ia,   verbose)
    _ecrire_onglet(ws_pm,   rows_pm,   verbose)
    _ecrire_onglet(ws_nore, rows_nore, verbose)

    wb.save(FICHIER)
