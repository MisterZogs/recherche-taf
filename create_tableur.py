import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Offres"

# ── Couleurs ──────────────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # bleu foncé
HEADER_FG   = "FFFFFF"
ROW_ODD     = "F2F7FF"
ROW_EVEN    = "FFFFFF"
GREEN_FILL  = "C6EFCE"   # postulé
YELLOW_FILL = "FFEB9C"   # en cours
ORANGE_FILL = "FCE4D6"   # à postuler priorité haute
GRAY_FILL   = "D9D9D9"   # hors profil

STAR_COLORS = {
    5: "FF0000",   # rouge – fit parfait
    4: "FF6600",   # orange – très bon
    3: "FFC000",   # jaune – moyen
    2: "92D050",   # vert clair – faible
    1: "BFBFBF",   # gris – hors profil
}

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── En-têtes ──────────────────────────────────────────────────────────────────
headers = [
    "Priorité", "Statut", "Poste", "Entreprise", "Source",
    "Contrat", "Localisation", "Remote", "Salaire / TJM",
    "Durée mission", "Fit / Notes", "Lien"
]
col_widths = [10, 12, 40, 22, 22, 12, 22, 10, 18, 14, 50, 50]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill    = make_fill(HEADER_BG)
    cell.font    = Font(color=HEADER_FG, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border  = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# ── Données ───────────────────────────────────────────────────────────────────
# Colonnes : Priorité(⭐), Statut, Poste, Entreprise, Source, Contrat,
#            Localisation, Remote, Salaire/TJM, Durée, Fit/Notes, Lien

rows = [
    # ── DÉJÀ DANS "À POSTULER" ─────────────────────────────────────────────
    (5, "Postulé",
     "Customer Success Manager AI",
     "Cominty",
     "Contact direct",
     "CDI", "Paris 9e", "Hybride", "60-75 K€", "",
     "Fit parfait — WallOfTraders.com (SaaS) + L'Oréal (onboarding/QBR) + IA générative",
     ""),

    (5, "À postuler",
     "Digital Process Owner – Recruiting & Onboarding",
     "Accor",
     "LinkedIn",
     "CDI", "Issy-les-Moulineaux", "Hybride", "n.c.", "",
     "Fit parfait — OneProfile/SuccessFactors L'Oréal colle directement",
     "https://www.linkedin.com/jobs/view/4426930800/"),

    (4, "À postuler",
     "People Operations Specialist",
     "Mercor",
     "LinkedIn",
     "Freelance", "100% Remote", "Full remote", "60-80 $/h", "",
     "Bon match — HRIS (SAP SF ✓, Workday, Oracle), Fortune 500 ✓, multi-pays ✓",
     "https://www.linkedin.com/jobs/view/4428131075/"),

    (3, "À postuler",
     "Talent Management Project Manager",
     "CHANEL",
     "LinkedIn",
     "CDD", "Paris / Londres", "Hybride", "n.c.", "",
     "Background HR OK — poste interne RH (pas CSM) et CDD",
     "https://www.linkedin.com/jobs/view/4416160999/"),

    (3, "À postuler",
     "Enablement Trainer",
     "Hostaway",
     "LinkedIn",
     "CDI", "100% Remote Europe", "Full remote", "n.c.", "",
     "Formation & coaching EMEA ✓ — loin du SAP/HR (hospitality SaaS)",
     "https://www.linkedin.com/jobs/view/4426695837/"),

    (2, "À postuler",
     "Senior Consultant HR Transformation Workday",
     "Deloitte",
     "LinkedIn",
     "CDI", "La Défense", "Hybride", "n.c.", "",
     "Consulting RH ✓ — certification Workday requise, stack différent de SAP",
     "https://www.linkedin.com/jobs/view/4416588225/"),

    (1, "Hors profil",
     "Data Consultant HCM / Workday",
     "Jobgether",
     "LinkedIn",
     "Freelance", "100% Remote", "Full remote", "n.c.", "",
     "SQL avancé, ETL, migration Workday — trop technique, mauvais stack",
     "https://www.linkedin.com/jobs/view/4425836016/"),

    # ── TROUVÉES LORS DE LA 1ère RECHERCHE ─────────────────────────────────
    (5, "À postuler",
     "SAP HCM Consultant – French Payroll",
     "Movement Group",
     "movementgroup.uk",
     "CDI", "Remote", "Full remote", "n.c.", "",
     "Fit parfait — SAP HCM paie française, remote",
     "https://movementgroup.uk/job/sap-hcm-consultant-time-management/"),

    (5, "À postuler",
     "Directeur de Programme SIRH Senior",
     "LeHibou",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "650-780 €/j", "18 mois",
     "Transformation RH globale, gouvernance, budget — profil senior idéal",
     "https://www.free-work.com/fr/tech-it/job-mission/administrateur-applicatif-erp-crm-sirh/mission-freelance-directeur-de-programme-sirh-senior-h-f"),

    (4, "À postuler",
     "Expert Intégration SAP SuccessFactors BTP",
     "LINKWAY",
     "free-work.com",
     "Freelance", "Luxembourg", "Full remote", "n.c.", "12 mois",
     "Projet bancaire, intégration SF + SAP BTP — full remote, anglais courant requis",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/expert-integration-sap-successfactors-btp"),

    (4, "À postuler",
     "Consultant SAP HR Time Management",
     "SAP-HIRE",
     "free-work.com",
     "Freelance", "La Garenne-Colombes", "Hybride 3j/2j", "600-700 €/j", "3 mois renouvelable",
     "Support SAP ECC HR TIME, configuration, incidents — 3j remote / 2j onsite",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-hr-time-management-1"),

    (4, "À postuler",
     "Customer Success Manager",
     "AssessFirst",
     "welcometothejungle.com",
     "CDI", "Paris", "Full remote", "n.c.", "",
     "HR Tech, full remote, renouvellement / adoption produit — profil CSM Senior",
     "https://www.welcometothejungle.com/fr/companies/assessfirst/jobs/customer-success-manager_paris_ASSES_M4x97o8"),

    # ── NOUVELLES OFFRES – FREE-WORK SAP HCM ───────────────────────────────
    (3, "À postuler",
     "Expert SAP HCM – Gestion des temps",
     "RED Commerce",
     "free-work.com",
     "Freelance", "La Garenne-Colombes", "n.p.", "n.c.", "3 mois",
     "Analyse besoins RH, configuration SAP HCM, formation utilisateurs",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/expert-sap-hcm-en-gestion-des-temps"),

    (3, "À postuler",
     "Consultant SAP ECC TIME",
     "EBMC",
     "free-work.com",
     "Freelance", "La Garenne-Colombes", "n.p.", "400-700 €/j", "6 mois",
     "Support fonctionnel GTA, configuration, incidents SAP ECC HCM",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-ecc-time"),

    (3, "À postuler",
     "Consultant SAP HR Paie & GTA",
     "KUBE Partners / Pixie Services",
     "free-work.com",
     "Freelance", "Île-de-France", "n.p.", "n.c.", "12 mois",
     "Pilotage fonctionnel RH + paie dans environnement SAP ECC6 fortement customisé",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-hr-paie-gta"),

    (3, "À postuler",
     "Consultant SAP HR/HCM (S/4HANA)",
     "EBMC",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "400-800 €/j", "6 mois",
     "Transformation S/4HANA, gestion des temps, ABAP, Fiori/UI5, déploiement international",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-hr-hcm-s4-hana"),

    (3, "À postuler",
     "Consultant SAP HR/HCM",
     "SAP-HIRE",
     "free-work.com",
     "Freelance", "Paris", "Hybride 3-4j onsite", "550-650 €/j", "6 mois",
     "Implémentation S/4 Core, gestion des temps, intégration Oracle HCM",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-hr-hcm-2"),

    (3, "À postuler",
     "Consultant SAP HR/HCM Time & Payroll S/4HANA",
     "DUONEXT",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "n.c.", "12 mois",
     "Migration agile S/4HANA, temps & paie, design → go-live → support",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-hr-hcm-time-payroll-s-4hana-h-f"),

    # ── NOUVELLES OFFRES – FREE-WORK SAP SUCCESSFACTORS ────────────────────
    (3, "À postuler",
     "Consultant SuccessFactors Employee Central",
     "Mindquest",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "n.c.", "6 mois",
     "Maintenance, upgrades, support utilisateurs RH & international, Story reports",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant/consultant-successfactors-ec-h-f-75"),

    (3, "À postuler",
     "Consultant SAP SuccessFactors – Anglais courant",
     "Intuition IT Solutions",
     "free-work.com",
     "Freelance", "Luxembourg", "Hybride 60% onsite", "400-580 €/j", "6 mois (juil-déc 2026)",
     "Release management, tous modules SF (profils, perf, succession, LMS)",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant-erp-ms-dynamics-oracle-sage-sap/consultant-sap-success-factor-anglais-courant-2-jours-tt"),

    (3, "À postuler",
     "Business Analyst Senior / Consultant Technico-Fonctionnel SaaS",
     "Bessand Freelances",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "400-530 €/j", "6 mois",
     "Configuration SaaS IA, lien client/équipes, ateliers fonctionnels, pré-prod",
     "https://www.free-work.com/fr/tech-it/job-mission/business-analyst/business-analyst-senior-consultant-technico-fonctionnel-saas"),

    (2, "À postuler",
     "PMO SuccessFactors",
     "Mindquest",
     "free-work.com",
     "Freelance", "Paris", "n.p.", "n.c.", "3 mois",
     "Coordination projets SF design/build/test, reporting, bilingue EN requis",
     "https://www.free-work.com/fr/tech-it/job-mission/project-management-officer/pmo-successfactors-h-f-75"),

    # ── NOUVELLES OFFRES – EURSAP ───────────────────────────────────────────
    (4, "À postuler",
     "SAP Consultant OR Trainer (Any module)",
     "Eursap",
     "eursap.eu",
     "CDI", "Remote", "Full remote", "80-100 K€ + equity", "",
     "French language, remote, consultant ou formateur SAP — bon fit profil formateur/consultant",
     "https://eursap.eu/sap-jobs/"),

    (2, "À postuler",
     "Global SAP HRIS Manager (SuccessFactors)",
     "Via Eursap",
     "eursap.eu",
     "CDI", "Allemagne", "Non précisé", "jusqu'à 145 K€", "",
     "Très bon salaire — poste en Allemagne, pas remote confirmé",
     "https://eursap.eu/jobs/global-sap-hris-manager-successfactors-34522-de"),

    # ── NOUVELLES OFFRES – INDEED / AUTRES ─────────────────────────────────
    (4, "À postuler",
     "Senior SAP SF Inhouse Consultant – Integration EC",
     "Hornbach Baumarkt AG",
     "indeed.fr",
     "CDI", "France", "Full remote", "n.c.", "",
     "Remote, avancement technique des systèmes RH via SuccessFactors EC",
     ""),

    (3, "À postuler",
     "Service Consultant SAP SuccessFactors",
     "delaware",
     "indeed.fr",
     "CDI", "France", "Non précisé", "n.c.", "",
     "3+ ans SIRH, implémentations SF end-to-end, cabinet conseil",
     ""),

    (3, "À postuler",
     "Lead Consultant SAP SuccessFactors Employee Central",
     "Tenth Revolution Group",
     "free-work.com / indeed.fr",
     "CDI", "Paris (75001)", "Hybride", "50-60 K€", "",
     "Pilotage implémentations SF EC bout en bout — salaire faible pour le profil senior",
     ""),

    # ── NOUVELLES OFFRES – FREELANCE-INFORMATIQUE ───────────────────────────
    (3, "À postuler",
     "Expert SIRH – GTA – SAP",
     "n.c.",
     "freelance-informatique.fr",
     "Freelance", "Paris (75)", "n.p.", "n.c.", "36 mois",
     "Mission longue, GTA + SAP — très longue durée, bonne visibilité",
     "https://www.freelance-informatique.fr/mission-expert-sirh-gta-sap-260507I001"),

    (3, "À postuler",
     "Consultant Senior SAP SuccessFactors – Recruiting & EC",
     "n.c.",
     "freelance-informatique.fr",
     "Freelance", "Luxembourg", "n.p.", "n.c.", "12 mois",
     "Modules Recruiting + Employee Central, anglais requis",
     "https://www.freelance-informatique.fr/mission-sap-hr-461"),

    (3, "À postuler",
     "Consultant AMOA SAP HR",
     "n.c.",
     "freelance-informatique.fr",
     "Freelance", "La Plaine Saint-Denis (93)", "n.p.", "n.c.", "6 mois",
     "MOA SAP HR, Île-de-France nord",
     "https://www.freelance-informatique.fr/mission-sap-hr-461"),

    (2, "À postuler",
     "Consultant fonctionnel SAP HR/HCM",
     "n.c.",
     "freelance-informatique.fr",
     "Freelance", "Clermont-Ferrand (63)", "n.p.", "n.c.", "9 mois",
     "Hors IDF — déplacement nécessaire",
     "https://www.freelance-informatique.fr/mission-sap-hr-461"),
]

# ── Remplissage ───────────────────────────────────────────────────────────────
STATUS_FILLS = {
    "Postulé":      make_fill("C6EFCE"),
    "Hors profil":  make_fill("D9D9D9"),
    "À postuler":   make_fill("FFFFFF"),
    "En cours":     make_fill("FFEB9C"),
    "Refusé":       make_fill("FCE4D6"),
}

for row_idx, data in enumerate(rows, start=2):
    stars, statut, poste, entreprise, source, contrat, loc, remote, salaire, duree, notes, lien = data

    row_fill = make_fill(ROW_ODD if row_idx % 2 == 0 else ROW_EVEN)

    for col_idx, value in enumerate([
        "⭐" * stars, statut, poste, entreprise, source,
        contrat, loc, remote, salaire, duree, notes, lien
    ], start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if col_idx in (1, 2, 6, 7, 8, 9, 10) else "left")

        # Couleur colonne Priorité
        if col_idx == 1:
            cell.fill = make_fill(STAR_COLORS.get(stars, "FFFFFF"))
            cell.font = Font(bold=True)
        # Couleur colonne Statut
        elif col_idx == 2:
            cell.fill = STATUS_FILLS.get(statut, row_fill)
        else:
            cell.fill = row_fill

    ws.row_dimensions[row_idx].height = 40

# ── Filtre auto ────────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# ── Légende ───────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Légende")
legends = [
    ("Priorité", ""),
    ("⭐⭐⭐⭐⭐", "Fit parfait"),
    ("⭐⭐⭐⭐",   "Très bon fit"),
    ("⭐⭐⭐",     "Fit moyen"),
    ("⭐⭐",       "Fit faible"),
    ("⭐",         "Hors profil"),
    ("", ""),
    ("Statut", ""),
    ("À postuler", "Pas encore envoyé"),
    ("Postulé",    "Candidature envoyée"),
    ("En cours",   "Entretien / échange en cours"),
    ("Refusé",     "Refus reçu"),
    ("Hors profil","Non pertinent"),
]
for r_idx, (k, v) in enumerate(legends, start=1):
    ws2.cell(row=r_idx, column=1, value=k).font = Font(bold=(v == ""))
    ws2.cell(row=r_idx, column=2, value=v)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 35

# ── Sauvegarde ────────────────────────────────────────────────────────────────
output = "/Users/gaetan/Documents/IA/recherche-taf/offres_emploi.xlsx"
wb.save(output)
print(f"Fichier créé : {output} — {len(rows)} offres")
