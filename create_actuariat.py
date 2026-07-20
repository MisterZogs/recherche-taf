"""
Crée offres_actuariat.xlsx — tableur offres d'emploi pour un ami actuaire.
Même structure que offres_emploi.xlsx (15 colonnes, feuilles Offres Actuariat + Fait).
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FICHIER = "/Users/gaetan/Documents/IA/recherche-taf/offres_actuariat.xlsx"

HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
ROW_ODD   = "F2F7FF"
ROW_EVEN  = "FFFFFF"

STAR_COLORS = {
    5: "00FF0000",
    4: "00FF8C00",
    3: "00FFD700",
    2: "0070AD47",
    1: "00969696",
}

COLS = [
    "Priorité", "Statut", "Fait", "Poste", "Entreprise", "Source", "Contrat",
    "Localisation", "Remote", "Salaire / TJM", "Durée mission",
    "Fit / Notes", "Lien", "CV à envoyer", "Prétention"
]

col_widths = [10, 12, 6, 40, 22, 22, 12, 22, 12, 18, 14, 50, 55, 20, 14]

CENTERED_COLS = {1, 2, 3, 6, 7, 8, 9, 10, 11, 14, 15}


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Offres ──────────────────────────────────────────────────────────────────
# (stars, statut, poste, entreprise, source, contrat, localisation, remote, salaire, duree, notes, lien)

rows = [
    # ── ⭐⭐⭐⭐⭐ Grands employeurs + full remote ──────────────────────────────
    (5, "À postuler",
     "Consultant(e) Senior Actuariat",
     "Sia Partners",
     "linkedin.com",
     "CDI", "Paris (75)", "Full remote", "n.c.", "",
     "Cabinet conseil, full remote confirmé — profil consultant actuariat senior",
     "https://fr.linkedin.com/jobs/view/consultant-e-senior-actuariat-at-sia-4431741276"),

    (5, "À postuler",
     "Consultant en actuariat F/H",
     "Onepoint",
     "linkedin.com",
     "CDI", "Paris (75)", "Full remote", "n.c.", "",
     "Cabinet conseil, full remote confirmé — profil consultant actuariat",
     "https://fr.linkedin.com/jobs/view/consultant-en-actuariat-f-h-at-onepoint-4350930215"),

    (5, "À postuler",
     "Portfolio Analytics and Forecasts Actuary",
     "Alan",
     "linkedin.com",
     "CDI", "Paris / Nice / Lyon / Marseille / Nantes...", "Full remote", "n.c.", "",
     "Insurtech full remote native, plusieurs villes FR — profil data + actuariat",
     "https://fr.linkedin.com/jobs/view/portfolio-analytics-and-forecasts-actuary-at-alan-4429563218"),

    (5, "À postuler",
     "Lead Pricing Actuary (French Market)",
     "SCOR",
     "actuarylist.com",
     "CDI", "Paris (75)", "n.c.", "n.c.", "",
     "Grand réassureur mondial — tarification marché FR, minimum Fellow + 5 ans",
     "https://www.actuarylist.com/actuarial-jobs/79583-scor"),

    (5, "À postuler",
     "Actuary P&C Reinsurance Confirmé H/F",
     "Aon",
     "actuarylist.com",
     "CDI", "Paris (75)", "n.c.", "n.c.", "",
     "Top courtier réassurance mondial — conseil DFA, modélisation, 4-7 ans exp.",
     "https://www.actuarylist.com/actuarial-jobs/39684-aon"),

    (5, "À postuler",
     "Corporate Actuarial Expert",
     "Allianz Partners",
     "actuarylist.com",
     "CDI", "Paris (75)", "n.c.", "n.c.", "",
     "Grand assureur mondial — réserves IFRS/IFRS17, P&C, min. 1 an actuariat",
     "https://www.actuarylist.com/actuarial-jobs/5377-allianz"),

    (5, "À postuler",
     "AXA Health Business – Life and Health Actuary F/M",
     "AXA",
     "linkedin.com",
     "CDI", "Puteaux (92)", "Hybride", "n.c.", "",
     "Grand assureur mondial — actuariat vie/santé international, anglais requis",
     "https://fr.linkedin.com/jobs/view/axa-health-business-life-and-health-actuary-f-m-at-axa-en-france-4429536717"),

    (5, "À postuler",
     "Actuaire Modélisation et Solvabilité Prévoyance Santé F/H",
     "AXA",
     "linkedin.com / welcometothejungle.com",
     "CDI", "Marseille (13)", "Hybride 3j/sem", "n.c.", "",
     "AXA France — modélisation Solvabilité II, prévoyance/santé, région Sud",
     "https://fr.linkedin.com/jobs/view/actuaire-mod%C3%A9lisation-et-solvabilit%C3%A9-pr%C3%A9voyance-sant%C3%A9-f-h-at-axa-en-france-4440994525"),

    (5, "À postuler",
     "Actuaire – Responsable Gestion Financière H/F",
     "Crédit Agricole Assurances",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Grand groupe bancassurance — ALM/gestion financière, poste responsable",
     "https://www.welcometothejungle.com/fr/companies/groupe-credit-agricole/jobs/actuaire-responsable-gestion-financiere-h-f_paris"),

    (5, "À postuler",
     "Directeur Actuariat Assurances de Dommages",
     "Société Générale",
     "indeed.fr",
     "CDI", "Courbevoie (92)", "n.c.", "n.c.", "",
     "Grande banque — direction actuariat IARD, poste senior/management",
     "https://fr.indeed.com/jobs?q=actuaire&l=France"),

    (5, "À postuler",
     "Actuaire – Retraite & avantages sociaux",
     "Société Générale",
     "indeed.fr",
     "CDI", "Paris (75)", "n.c.", "n.c.", "",
     "Grande banque — actuariat retraite et avantages sociaux salariés",
     "https://fr.indeed.com/jobs?q=actuaire&l=France"),

    (5, "À postuler",
     "Consultant(e) Actuaire Senior",
     "Actuelia",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Cabinet actuariat pur-player — consulting assurance, profil senior, très bon fit",
     "https://www.welcometothejungle.com/fr/companies/actuelia/jobs/consultant-actuaire-senior"),

    # ── ⭐⭐⭐⭐ Cabinets conseil + freelance avec TJM ──────────────────────────
    (4, "À postuler",
     "Consultant(e) Actuaire Junior",
     "Actuelia",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Cabinet actuariat pur-player — profil junior/débutant, très bon fit cabinet spécialisé",
     "https://www.welcometothejungle.com/fr/companies/actuelia/jobs/actuaire-junior_paris"),

    (4, "À postuler",
     "Consultant Actuariat F/H",
     "Deloitte",
     "linkedin.com",
     "CDI", "La Défense (92)", "Hybride", "n.c.", "",
     "Big 4 — consulting actuariat toutes branches, profil junior à senior",
     "https://fr.linkedin.com/jobs/view/consultant-actuariat-f-h-at-deloitte-4075288008"),

    (4, "À postuler",
     "Consultant Manager Actuariat Vie",
     "KPMG",
     "indeed.fr",
     "CDI", "Courbevoie (92)", "n.c.", "n.c.", "",
     "Big 4 — management actuariat vie, client assureurs grands comptes",
     "https://fr.indeed.com/jobs?q=actuaire&l=France"),

    (4, "À postuler",
     "Consultant Senior Actuariat Assurance IARD",
     "KPMG",
     "indeed.fr",
     "CDI", "Courbevoie (92)", "n.c.", "n.c.", "",
     "Big 4 — conseil actuariat IARD/non-vie, 4+ ans exp., expertise technique",
     "https://fr.indeed.com/jobs?q=actuaire&l=France"),

    (4, "À postuler",
     "Consultant Senior Actuariat Compensation & Benefits H/F",
     "PwC France",
     "hellowork.com",
     "CDI", "Neuilly-sur-Seine (92)", "Hybride", "n.c.", "",
     "Big 4 — actuariat C&B (retraite, prévoyance, stock-options), IAS 19",
     "https://www.hellowork.com/fr-fr/emplois/81228158.html"),

    (4, "À postuler",
     "PI Tech & Cyber Underwriter",
     "Oliver James Associates",
     "theactuaryjobs.com",
     "CDI", "Paris (75)", "Hybride 2j/sem", "85 000 – 135 000 €/an + var.", "",
     "Grand assureur — souscription PI/Tech/Cyber upper market, bilingue FR/EN",
     "https://www.theactuaryjobs.com/job/176670/pi-tech-and-cyber-underwriter/"),

    (4, "À postuler",
     "Actuaire ALM & Modélisation – Épargne Retraite S2",
     "Oliver James Associates",
     "theactuaryjobs.com",
     "CDI", "Paris (75)", "n.c.", "~80 000 €/an", "",
     "Assureur vie international — ALM épargne/retraite Solvabilité II, 7+ ans",
     "https://www.theactuaryjobs.com/job/176917/actuaire-alm-and-modelisation-epargne-retraite-s2-/"),

    (4, "À postuler",
     "Pricing Lead (Reinsurance/Europe)",
     "Orange Malone",
     "theactuaryjobs.com",
     "CDI", "Paris / EU (multi-localisations)", "n.c.", "Salaire attractif + bonus", "",
     "Grand réassureur — lead pricing EMEA, Paris possible parmi autres villes EU",
     "https://www.theactuaryjobs.com/job/176866/pricing-lead-reinsurance-europe-/"),

    (4, "À postuler",
     "Actuaire / Ingénieur ALM Senior H/F",
     "FINAX Consulting",
     "free-work.com",
     "Mission / Freelance", "Paris (75)", "Hybride", "n.c.", "Démarrage sept. 2026",
     "Cabinet conseil actuariat — mission ALM senior, Paris, télétravail partiel",
     "https://www.free-work.com/fr/tech-it/job-mission/ingenieur-apres-vente/actuaire-ingenieur-alm-senior-h-f"),

    (4, "À postuler",
     "Consultant Actuaire Solvabilité II – ORSA & Gestion du Capital H/F",
     "FINAX Consulting",
     "free-work.com",
     "Mission / Freelance", "Rouen (76)", "Hybride 3j/sem TT", "400 – 800 €/j", "5 mois (juil.–déc. 2026)",
     "ORSA, capital économique, Solvabilité II, Normandie avec télétravail majoritaire",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant/consultant-actuaire-solvabilite-ii-orsa-gestion-du-capital-h-f"),

    (4, "À postuler",
     "Actuaire Senior Santé & Prévoyance",
     "Valayo",
     "linkedin.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Cabinet actuariat — profil senior santé/prévoyance, missions clients assureurs",
     "https://fr.linkedin.com/jobs/view/actuaire-senior-sant%C3%A9-pr%C3%A9voyance-at-valayo-4440064375"),

    (4, "À postuler",
     "Actuaire Vie Expérimenté F/H",
     "Taylor Made Recrutement",
     "linkedin.com",
     "CDI", "Paris (75)", "Hybride", "75 000 – 85 000 €/an", "",
     "Cabinet recrutement spécialisé assurance — profil actuaire vie expérimenté",
     "https://fr.linkedin.com/jobs/view/actuaire-vie-exp%C3%A9riment%C3%A9-f-h-cdi-paris-75-85-k%C3%AC-at-taylor-made-recrutement-4434851370"),

    (4, "À postuler",
     "Talent Pool Actuaire – Pricing / Profitability",
     "Seyna",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "45 000 – 105 000 €/an", "",
     "Insurtech innovante — tarification produits assurance, profil junior à senior",
     "https://www.welcometothejungle.com/fr/companies/seyna/jobs/talent-pool-actuaire-pricing-profitability_paris"),

    (4, "À postuler",
     "Consultant Senior Actuariat Santé Prévoyance B2B H/F",
     "Aops Conseil",
     "hellowork.com",
     "CDI", "Paris (75)", "n.c.", "90 000 €/an", "",
     "Cabinet actuariat — santé/prévoyance B2B, bon salaire affiché, profil senior",
     "https://www.hellowork.com/fr-fr/emplois/78025126.html"),

    (4, "À postuler",
     "Actuaire Produit Épargne – Chargé d'études",
     "Ufirst Advisory",
     "hellowork.com",
     "CDI", "Puteaux (92)", "n.c.", "70 000 – 90 000 €/an", "",
     "Conseil actuariat — épargne/retraite, fourchette salariale attractive",
     "https://www.hellowork.com/fr-fr/emplois/80401572.html"),

    # ── ⭐⭐⭐ Assureurs mid-range + freelance province ─────────────────────────
    (3, "À postuler",
     "Consultant Actuaire Inventaire Non-Vie Senior",
     "FINAX Consulting",
     "free-work.com",
     "Mission / Freelance", "Rouen (76)", "Hybride 1-2j/sem TT", "400 – 800 €/j", "4 mois (juin–oct. 2026)",
     "Inventaire non-vie, Normandie, télétravail limité — TJM attractif",
     "https://www.free-work.com/fr/tech-it/job-mission/consultant/consultant-actuaire-inventaire-non-vie-senior"),

    (3, "À postuler",
     "Actuaire Indépendant – Tarification Gamme Frais de Santé",
     "confidentiel",
     "freelance-informatique.fr",
     "Mission / Freelance", "Full remote", "Full remote", "n.c.", "3 jours ouvrés",
     "Mission courte full remote — tarification frais de santé, profil indépendant",
     "https://www.freelance-informatique.fr/consultant-actuaire-freelance-n1059"),

    (3, "À postuler",
     "Actuaire Prévoyance Assurance Vie – Audit Interne",
     "confidentiel",
     "freelance-informatique.fr",
     "Mission / Freelance", "Paris (75)", "n.c.", "n.c.", "3 mois",
     "Audit interne prévoyance/vie, Paris, 3 mois",
     "https://www.freelance-informatique.fr/consultant-actuaire-freelance-n1059"),

    (3, "À postuler",
     "Consultant Risk / Finance / Actuaire",
     "confidentiel",
     "freelance-informatique.fr",
     "Mission / Freelance", "Paris (75)", "n.c.", "n.c.", "3 mois",
     "Mission risk/finance/actuariat Paris — profil polyvalent",
     "https://www.freelance-informatique.fr/consultant-actuaire-freelance-n1059"),

    (3, "À postuler",
     "Actuaire Tarification Assurance B2B",
     "confidentiel",
     "babyloneconsulting.fr",
     "Mission / Freelance", "Nouvelle-Aquitaine", "n.c.", "n.c.", "",
     "Mission région Nouvelle-Aquitaine (proche Anglet) — tarification B2B",
     "https://www.babyloneconsulting.fr/mission-freelance/freelance-actuaire-tarification-assurance-b2b/"),

    (3, "À postuler",
     "Actuaire santé H/F",
     "Carte Blanche Partenaires",
     "linkedin.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Réseau soins de santé — actuariat produits santé, profil junior à confirmé",
     "https://fr.linkedin.com/jobs/view/actuaire-sant%C3%A9-h-f-at-carte-blanche-partenaires-4441201220"),

    (3, "À postuler",
     "Consultant Actuaire Senior",
     "CORRELIUM",
     "linkedin.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Cabinet conseil actuariat — profil senior, missions clients assureurs",
     "https://fr.linkedin.com/jobs/view/consultant-actuaire-senior-at-correlium-4439008626"),

    (3, "À postuler",
     "Actuaire Gestion du Capital H/F",
     "Abeille Assurances",
     "linkedin.com",
     "CDI", "Bois-Colombes (92)", "Hybride", "n.c.", "",
     "Groupe Abeille (ex-Aviva France) — gestion du capital Solvabilité II",
     "https://fr.linkedin.com/jobs/view/actuaire-gestion-du-capital-h-f-at-abeille-assurances-4430997173"),

    (3, "À postuler",
     "Actuaire / Data Scientist Senior",
     "Abeille Assurances",
     "linkedin.com",
     "CDI", "Bois-Colombes (92)", "Hybride", "n.c.", "",
     "Groupe Abeille — double compétence actuariat + data science, profil hybride",
     "https://fr.linkedin.com/jobs/view/actuaire-data-scientist-senior-at-abeille-assurances-4438684279"),

    (3, "À postuler",
     "Actuaire Confirmé – Tarification Emprunteur Individuel H/F",
     "CNP Assurances",
     "welcometothejungle.com",
     "CDI", "Issy-les-Moulineaux (92)", "Hybride", "n.c.", "",
     "Grand assureur — tarification crédit/emprunteur, profil confirmé, télétravail fréquent",
     "https://www.welcometothejungle.com/fr/companies/cnp-assurances/jobs/actuaire-confirme-tarification-emprunteur-individuel-h-f_issy-les-moulineaux"),

    (3, "À postuler",
     "Actuaire H/F",
     "CNP Assurances",
     "indeed.fr",
     "CDI", "Paris (75)", "Hybride", "60 000 – 80 000 €/an", "",
     "Grand assureur — profil généraliste, fourchette salariale claire",
     "https://fr.indeed.com/jobs?q=actuaire&l=France"),

    (3, "À postuler",
     "Actuaire F/H/X",
     "Garance (mutuelle)",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "50 – 58 €/h", "",
     "Mutuelle santé/prévoyance — taux horaire affiché, profil généraliste",
     "https://www.welcometothejungle.com/fr/companies/garance-mutuelle/jobs/actuaire-f-h-x_paris_GARAN_5mmo3GK"),

    (3, "À postuler",
     "Actuaire Consultant",
     "PRO BTP Groupe",
     "linkedin.com",
     "CDI", "Vanves (92)", "Hybride", "n.c.", "",
     "Institution prévoyance BTP — profil consultant actuariat interne",
     "https://fr.linkedin.com/jobs/view/actuaire-consultant-at-pro-btp-groupe-4437996941"),

    (3, "À postuler",
     "Consultant en Actuariat Junior H/F",
     "RSM France",
     "linkedin.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "Cabinet audit/conseil mid-market — actuariat junior, profil débutant bienvenu",
     "https://fr.linkedin.com/jobs/view/consultant-en-actuariat-junior-h-f-at-rsm-france-4438272243"),

    (3, "À postuler",
     "Actuaire ALM H/F",
     "Groupe VYV",
     "welcometothejungle.com",
     "CDI", "Paris (75)", "Hybride", "n.c.", "",
     "1er groupe mutualiste FR — ALM groupe VYV (MGEN, MNT, Harmonie Mutuelle)",
     "https://www.welcometothejungle.com/fr/companies/groupe-vyv/jobs/gestionnaire-d-actifs-h-f_paris_GV_yxabVy8"),

    (3, "À postuler",
     "Actuaire Modèles Vie F/H",
     "Morgan Philips Executive Search",
     "linkedin.com",
     "CDI", "Courbevoie (92)", "Hybride", "n.c.", "",
     "Via cabinet recrutement — modèles actuariels vie (Prophet probable), IDF",
     "https://fr.linkedin.com/jobs/view/actuaire-mod%C3%A8les-vie-f-h-fr877727-at-morgan-philips-executive-search-4431856305"),

    (3, "À postuler",
     "Actuaire Modélisation Vie (Prophet) H/F",
     "Fed Finance",
     "hellowork.com",
     "CDI", "Paris (75)", "n.c.", "65 000 €/an", "",
     "Via cabinet recrutement — modélisation vie/Prophet, salaire affiché",
     "https://www.hellowork.com/fr-fr/emplois/80404826.html"),

    # ── ⭐⭐ Province / présentiel / petite structure ─────────────────────────
    (2, "À postuler",
     "Actuaire Senior H/F",
     "SKILLS Banque et Assurance",
     "hellowork.com",
     "CDI", "Lyon (69)", "n.c.", "65 000 – 80 000 €/an", "",
     "Via cabinet — région lyonnaise, bon salaire mais remote non confirmé",
     "https://www.hellowork.com/fr-fr/emplois/81285128.html"),

    (2, "À postuler",
     "Manager Actuariat & Réassurance",
     "IMA – Inter Mutuelles Assistance",
     "hellowork.com",
     "CDI", "Niort (79)", "Présentiel", "80 000 – 110 000 €/an", "",
     "Très bon salaire mais Niort présentiel — poste management actuariat/réassurance",
     "https://www.hellowork.com/fr-fr/emplois/80420124.html"),

    (2, "À postuler",
     "Actuaire H/F",
     "Zurich Insurance",
     "linkedin.com",
     "CDI", "Paris (75)", "Présentiel", "n.c.", "",
     "Grand assureur international — présentiel imposé, Paris, profil généraliste",
     "https://fr.linkedin.com/jobs/view/actuaire-h-f-at-zurich-insurance-4435471934"),

    (2, "À postuler",
     "Actuaire Vie H/F",
     "Suravenir",
     "hellowork.com",
     "CDI", "Brest (29)", "n.c.", "n.c.", "",
     "Filiale Crédit Mutuel Bretagne — Brest, remote non indiqué",
     "https://www.hellowork.com/fr-fr/emplois/actuaire.html"),

    (2, "À postuler",
     "Actuaire Tarification & Rentabilité F/H",
     "Groupe BPCE",
     "linkedin.com",
     "CDI", "Saint-Grégoire (35)", "Hybride", "n.c.", "",
     "Groupe bancassurance — Rennes/St-Grégoire, hybride mais province",
     "https://fr.linkedin.com/jobs/view/actuaire-tarification-amp-rentabilit%C3%A9-f-h-rennes-at-groupe-bpce-4439245997"),

    (2, "À postuler",
     "Actuaire Risques H/F",
     "Thélem Assurances",
     "indeed.fr / hellowork.com",
     "CDI", "Chécy (45)", "Présentiel", "50 000 – 70 000 €/an", "",
     "Assureur régional Centre-Val-de-Loire — présentiel, salaire correct",
     "https://www.hellowork.com/fr-fr/emplois/62627335.html"),

    (2, "À postuler",
     "Actuaire Prévoyance H/F",
     "LHH Recruitment Solutions",
     "hellowork.com",
     "CDI", "Strasbourg (67)", "n.c.", "45 000 – 60 000 €/an", "",
     "Via cabinet — Bas-Rhin (Alsace), remote non confirmé, profil junior/confirmé",
     "https://www.hellowork.com/fr-fr/emplois/79124000.html"),

    # ── ⭐ Junior / stage / hors profil actuarial pur ──────────────────────────
    (1, "À postuler",
     "Junior Assistant Underwriter",
     "Oliver James Associates",
     "theactuaryjobs.com",
     "CDD", "Paris (75)", "n.c.", "40 000 – 50 000 €/an", "",
     "Poste underwriting junior (pas actuariat pur), CDD — à garder en backup",
     "https://www.theactuaryjobs.com/job/176754/junior-assistant-underwriter/"),

    (1, "À postuler",
     "Actuarial Apprentice",
     "SCOR",
     "actuarylist.com",
     "Apprentissage", "Paris (75)", "n.c.", "n.c.", "",
     "Alternance SCOR — modélisation risques médicaux réassurance, Python/Java requis",
     "https://www.actuarylist.com/actuarial-jobs/77857-scor"),
]

# ── Construction du workbook ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Offres Actuariat"


def write_headers(ws_target):
    for col_idx, (h, w) in enumerate(zip(COLS, col_widths), start=1):
        cell = ws_target.cell(row=1, column=col_idx, value=h)
        cell.fill = make_fill(HEADER_BG)
        cell.font = Font(color=HEADER_FG, bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws_target.column_dimensions[get_column_letter(col_idx)].width = w
    ws_target.row_dimensions[1].height = 30
    ws_target.freeze_panes = "A2"


write_headers(ws)

# Tri par priorité décroissante
rows.sort(key=lambda r: -r[0])

for row_idx, data in enumerate(rows, start=2):
    stars, statut, poste, entreprise, source, contrat, loc, remote, salaire, duree, notes, lien = data
    prio_val = "⭐" * stars

    row_fill = make_fill(ROW_ODD if row_idx % 2 == 0 else ROW_EVEN)

    values = [prio_val, statut, "", poste, entreprise, source, contrat, loc, remote,
              salaire, duree, notes, lien, "", ""]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border()
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
            horizontal="center" if col_idx in CENTERED_COLS else "left"
        )
        if col_idx == 1:
            cell.fill = make_fill(STAR_COLORS.get(stars, "FFFFFF"))
            cell.font = Font(bold=True)
        else:
            cell.fill = row_fill

    ws.row_dimensions[row_idx].height = 40

ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

# ── Feuille Fait (archive) ───────────────────────────────────────────────────
ws_fait = wb.create_sheet("Fait")
write_headers(ws_fait)

wb.save(FICHIER)
print(f"Créé : {FICHIER} — {len(rows)} offres")
