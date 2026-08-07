"""Repare 5 lignes de l'onglet 'Offres SIRH' dont les colonnes etaient decalees.

Les valeurs partaient de la colonne B (Statut) au lieu de la colonne D (Poste),
laissant Statut et Fait remplis par le titre et l'entreprise, et la colonne Lien
remplie par une date. Le remappage est explicite pour chaque ligne, cle sur le
titre tel qu'il se trouve actuellement dans la colonne Statut.
"""

import openpyxl
from openpyxl.styles import PatternFill

FICHIER = 'offres_emploi.xlsx'

FILLS = {
    '⭐⭐⭐⭐⭐': 'FFFF0000',
    '⭐⭐⭐⭐':   'FFFF8C00',
    '⭐⭐⭐':     'FFFFD700',
    '⭐⭐':       'FF70AD47',
    '⭐':         'FF808080',
}

N = 'https://www.hellowork.com/fr-fr/emplois/81222357.html'

# titre (valeur actuellement en colonne Statut) -> ligne complete corrigee (17 valeurs)
CORRECTIONS = {
    'Consultant Senior SIRH SAP SF Performance & Goals H/F': [
        '⭐⭐⭐⭐', None, None,
        'Consultant Senior SIRH SAP SF Performance & Goals H/F',
        'Grand groupe international', 'Hellowork', 'Freelance', 'Paris (75)',
        'Hybride', 'NC', 'NC',
        'SF PMGM, 10 ans requis, gouvernance SIRH, pilotage roadmap. Fit fort sur module Performance & Goals.',
        N, 'Resume_GaetanFRANCOIS_SIRH.pdf', '650-750€/j', '11/07/2026', '11/07/2026',
    ],
    'Consultant SIRH SuccessFactors (9 mois)': [
        '⭐⭐⭐', None, None,
        'Consultant SIRH SuccessFactors',
        'Non communiqué', 'freelance-informatique.fr', 'Freelance',
        'Issy-les-Moulineaux (92)', 'NC', 'NC', '9 mois',
        'Mission 9 mois SAP SuccessFactors, Issy-les-Moulineaux. Profil Gaëtan SF multi-modules = aligné.',
        'https://www.freelance-informatique.fr/mission-consultant-technique-sirh-sap-successfactors-n16466',
        'Resume_GaetanFRANCOIS_SIRH.pdf', '650-750€/j', '07/2026', 'NC',
    ],
    'Manager SIRH - Nantes (CDI)': [
        '⭐⭐⭐', None, None,
        'Manager SIRH',
        'ALTHEA', 'Welcome to the Jungle', 'CDI', 'Saint-Herblain / Nantes',
        'Hybride', '48-62K€', 'NC',
        'Cabinet conseil SIRH, pilotage projets de A à Z, conduite du changement. Nantes = frein pour remote Anglet.',
        'https://www.welcometothejungle.com/fr/companies/althea/jobs/manager-sirh-nantes-cdi-h-f_saint-herblain',
        'Resume_GaetanFRANCOIS_SIRH.pdf', '48-62K€', '07/2026', 'NC',
    ],
    'SAP Payroll Implementation Lead - France': [
        '⭐⭐⭐', None, None,
        'SAP Payroll Implementation Lead - France',
        'Strada Global', 'Strada Global Careers / Remote Rocketship', 'CDI', 'France',
        '100% remote', 'NC', 'NC',
        'Strada = RH/paie globale (SAP, Oracle, Workday). Lead impl SAP Payroll France, 5+ implémentations SAP HCM requises, connaissance paie FR. Profil Gaëtan SAP HR adjacent mais payroll pur = frein partiel.',
        'https://careers.stradaglobal.com/careers/job/1133913392690-sap-payroll-implementation-lead-france-granada-spain',
        'Resume_GaetanFRANCOIS_SIRH.pdf', 'NC', '07/2026', 'NC',
    ],
    'Senior Manager SIRH - Lyon (CDI)': [
        '⭐⭐', None, None,
        'Senior Manager SIRH',
        'ALTHEA', 'Welcome to the Jungle', 'CDI', 'Lyon',
        'Hybride', 'NC', 'NC',
        'ALTHEA cabinet conseil SIRH. Senior Manager, Lyon. Frein : Lyon pas remote. Profil senior bien aligné sinon.',
        'https://www.welcometothejungle.com/fr/companies/althea/jobs/senior-manager-sirh-lyon-cdi_lyon',
        'Resume_GaetanFRANCOIS_SIRH.pdf', 'NC', '07/2026', 'NC',
    ],
}


CORRECTIONS_CSM = {
    'Senior Customer Success Manager': [
        '⭐⭐⭐⭐', None, None,
        'Senior Customer Success Manager',
        'Sprinklr', 'Jobgether / Remote Rocketship', 'CDI', 'France',
        '100% remote', 'NC', 'NC',
        'Plateforme CX unifiée (social, service, marketing). CSM stratégique sur comptes enterprise. Remote France.',
        'https://jobgether.com/offer/6a0e8a64cc7b72676990a7dc-senior-customer-success-manager',
        'Resume_GaetanFRANCOIS_CSM_EN.pdf', 'NC', '05/2026', '05/2026',
    ],
    'Customer Success Manager': [
        '⭐⭐', None, None,
        'Customer Success Manager',
        'Okta', 'Welcome to the Jungle', 'CDI', 'Paris',
        'Hybride', '71-98K€ OTE', 'NC',
        'IAM/SSO, CSM enterprise, OTE 71-98K€. FREIN MAJEUR : français + italien requis (pas de mention anglais seul).',
        'https://www.welcometothejungle.com/fr/companies/okta/jobs/customer-success-manager_paris_mnpmts4t',
        'Resume_GaetanFRANCOIS_CSM_EN.pdf', '71-98K€', '07/2026', 'NC',
    ],
}

# L'ancienne ligne IA porte la meme mission que celle ajoutee le 07/08/2026 : le lien
# etait range dans la colonne Fit/Notes, donc invisible pour la deduplication. On garde
# une seule ligne, avec l'URL canonique et les notes les plus completes.
DOUBLON_IA_ANCIEN = 'Consultant IA Générative & Conduite du changement'
DOUBLON_IA_LIEN_NEUF = 'https://www.mission-freelances.fr/missions/consultant-ia-generative-conduite-du-changement-paris-87a17202/'


def _reparer(ws, corrections):
    """Idempotent : une ligne deja reparee (titre en colonne Poste) est ignoree."""
    restantes = dict(corrections)
    n = 0
    for row in ws.iter_rows(min_row=2):
        cle = str(row[1].value)
        if cle not in restantes:
            continue
        valeurs = restantes.pop(cle)
        for cell, val in zip(row, valeurs):
            cell.value = val
        row[0].fill = PatternFill(fill_type='solid', fgColor=FILLS[valeurs[0]])
        print(f"{ws.title} ligne {row[0].row} reparee : {valeurs[3]}")
        n += 1

    postes = {str(r[3].value) for r in ws.iter_rows(min_row=2)}
    for cle, valeurs in restantes.items():
        if valeurs[3] not in postes:
            raise SystemExit(f"{ws.title} : ligne introuvable et non deja reparee -> {cle}")
        print(f"{ws.title} : deja reparee, ignoree -> {valeurs[3]}")
    return n


def _fusionner_doublon_ia(ws):
    """Supprime l'ancienne ligne mal formee, la version ajoutee le 07/08 la remplace."""
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == DOUBLON_IA_ANCIEN:
            cible = next(r for r in ws.iter_rows(min_row=2)
                         if str(r[12].value) == DOUBLON_IA_LIEN_NEUF)
            print(f"{ws.title} ligne {row[0].row} : doublon de la ligne {cible[0].row}, fusionnee")
            ws.delete_rows(row[0].row)
            return 1
    raise SystemExit('doublon IA introuvable')


def main():
    wb = openpyxl.load_workbook(FICHIER)
    n = _reparer(wb['Offres SIRH'], CORRECTIONS)
    n += _reparer(wb['Offres CSM'], CORRECTIONS_CSM)
    n += _fusionner_doublon_ia(wb['Offres IA'])
    wb.save(FICHIER)
    print(f"{n} lignes traitees")


if __name__ == '__main__':
    main()
