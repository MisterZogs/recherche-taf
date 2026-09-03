#!/usr/bin/env python3
"""
Relance de recherche du 02/09/2026 — fusion des 4 clusters parallèles.

Ordre d'insertion : cluster 2 (porteur des marqueurs 'Offres USA') en premier,
puis Pays Basque, puis remote/IA/PM, puis FR/freelance. Ainsi une offre trouvée
à la fois avec et sans le marqueur USA est routée vers 'Offres USA'.

Dédoublonnage en deux couches :
  1. ici, sur lien NORMALISÉ (segment de locale Workday, utm_*, /apply, /application,
     slash final) — les deux clusters ATS et Pays Basque ont constaté que sans ça
     tous les postes Workday déjà en base repassent pour neufs ;
  2. dans add_offre.ajouter_offres(), sur lien brut (garde-fou historique).
"""
import importlib.util
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import add_offre

SCRATCH = Path('/private/tmp/claude-501/-Users-gaetan-Documents-IA-recherche-taf'
               '/11bfea0a-0a04-4347-98ff-9a9b072e379c/scratchpad')

FICHIERS = [
    'offres_cluster2_ats.py',
    'offres_cluster4_paysbasque.py',
    'offres_cluster3_remote.py',
    'offres_cluster1_fr.py',
]


def normaliser(url: str) -> str:
    """Ramène deux écritures d'une même offre à la même clé."""
    u = (url or '').strip()
    if not u:
        return ''
    u = u.split('?utm_')[0].split('&utm_')[0]
    # Workday : /en-US/, /fr-FR/ entre le domaine et le site ID
    u = re.sub(r'(myworkdayjobs\.com)/[a-z]{2}-[A-Z]{2}/', r'\1/', u)
    u = re.sub(r'/(application|apply)/?$', '', u)
    return u.rstrip('/').lower()


def charger(nom: str) -> list[dict]:
    spec = importlib.util.spec_from_file_location('m', SCRATCH / nom)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.OFFRES


def liens_en_base() -> set[str]:
    wb = openpyxl.load_workbook(add_offre.FICHIER)
    liens = set()
    for ws in wb.worksheets:
        hdr = [c.value for c in ws[1]]
        if 'Lien' not in hdr:
            continue
        i = hdr.index('Lien')
        for row in ws.iter_rows(min_row=2):
            v = row[i].value
            if v:
                liens.add(normaliser(str(v)))
    return liens


def main():
    connus = liens_en_base()
    print(f"{len(connus)} liens normalisés déjà en base")

    offres, vus = [], set()
    for nom in FICHIERS:
        brut = charger(nom)
        garde = 0
        for o in brut:
            cle = normaliser(str(o.get('Lien') or ''))
            if not cle:
                continue
            if cle in connus or cle in vus:
                continue
            vus.add(cle)
            offres.append(o)
            garde += 1
        print(f"{nom:35s} {len(brut):4d} candidates → {garde:4d} retenues")

    print(f"\n{len(offres)} offres à insérer\n")
    add_offre.ajouter_offres(offres)


if __name__ == '__main__':
    main()
