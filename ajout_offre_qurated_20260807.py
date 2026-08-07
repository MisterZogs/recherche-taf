import openpyxl
from copy import copy

wb = openpyxl.load_workbook('offres_emploi.xlsx')
ws = wb['Fait']

modele = ws[2]  # ligne modele deja stylee (priorite 5 etoiles, rouge)
r = ws.max_row + 1

valeurs = [
    '⭐⭐⭐⭐⭐',
    'Postulé',
    'x',
    'Global Product Owner - SuccessFactors',
    'Qurated (pour un cabinet d avocats international)',
    'LinkedIn',
    'CDD / FTC 18 mois',
    'Royaume-Uni',
    '100% remote',
    None,
    '18 mois',
    "Ownership unique de la plateforme SAP SuccessFactors sur tout le perimetre mondial : vision et roadmap, solution design multi-modules (Employee Central, Performance & Goals, Compensation, Recruiting, Onboarding, Learning), standards globaux, cycles de release, conduite du changement et relation editeur. Fit tres eleve : SuccessFactors multi-modules, contexte global et conduite du changement correspondent au parcours L'Oreal. Full remote. Postule le 07/08/2026.",
    'https://www.linkedin.com/jobs/view/4447393087',
    'Resume_GaetanFRANCOIS_SIRH_EN.pdf',
    None,
    '07/08/2026',
    '06/08/2026',
]

for i, val in enumerate(valeurs, start=1):
    cell = ws.cell(row=r, column=i, value=val)
    src = modele[i - 1]
    cell.fill = copy(src.fill)
    cell.font = copy(src.font)
    cell.alignment = copy(src.alignment)
    cell.border = copy(src.border)

ws.row_dimensions[r].height = ws.row_dimensions[2].height

wb.save('offres_emploi.xlsx')
print('Ligne ajoutee en', r)
